import openpyxl

def inspect_sheet(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    print(f"Inspecting sheet: {sheet_name}")
    for i, row in enumerate(sheet.iter_rows(values_only=False)):
        vals = [str(c.value) if c.value is not None else "" for c in row]
        if any(vals):
            # Check for "RESPOSTES"
            if any("RESPOSTES" in v.upper() for v in vals):
                print(f"--- FOUND RESPOSTES AT ROW {i+1} ---")
                print(vals)
            # Check for bold in options
            bold = [c.font.bold if c.font else False for c in row]
            if any(bold):
                 # print(f"Row {i+1} has BOLD: {vals}")
                 pass
            # Check for background color
            fill = [c.fill.start_color.index if c.fill and c.fill.start_color and c.fill.start_color.index != '00000000' else None for c in row]
            if any(fill):
                # print(f"Row {i+1} has FILL: {fill} | {vals}")
                pass
    
    # Print last 20 rows to see if answers are there
    last_rows = list(sheet.iter_rows(values_only=True))[-20:]
    print("\n--- LAST 20 ROWS ---")
    for r in last_rows:
        print(r)

if __name__ == "__main__":
    inspect_sheet('PREGUNTES.xlsx', 'T8 FÍSICA')
