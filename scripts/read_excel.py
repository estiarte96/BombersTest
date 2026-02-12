import openpyxl
import sys

def read_excel(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        print(f"Sheets: {workbook.sheetnames}")
        
        for name in workbook.sheetnames:
            sheet = workbook[name]
            print(f"\n--- Sheet: {name} ---")
            count = 0
            for row in sheet.iter_rows(values_only=True):
                if any(row):  # Skip empty rows
                    print("\t".join([str(cell).replace('\n', ' ') if cell is not None else "" for cell in row]))
                    count += 1
                if count >= 30: # Only first 30 rows per sheet for overview
                    print("... (truncated)")
                    break
            
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        read_excel(sys.argv[1])
    else:
        print("Please provide a file path.")
