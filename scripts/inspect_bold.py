import openpyxl

def inspect_bold(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found.")
        return
    
    sheet = wb[sheet_name]
    print(f"Inspecting sheet: {sheet_name}")
    for i, row in enumerate(sheet.iter_rows(values_only=False)):
        for cell in row:
            if cell.value and cell.font and cell.font.bold:
                # Check if it looks like an option (a), b), etc)
                val = str(cell.value).strip()
                if any(val.startswith(p) for p in ['a)', 'b)', 'c)', 'd)', 'a.', 'b.', 'c.', 'd.']):
                     print(f"Row {i+1}, Col {cell.column_letter}: [BOLD] {val}")
        if i > 100: # Limit output
            break

if __name__ == "__main__":
    inspect_bold('PREGUNTES.xlsx', '1 CE')
