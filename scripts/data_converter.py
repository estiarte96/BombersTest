import openpyxl
import json
import re

def parse_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    data = []
    
    skip_sheets = ['TOTS ELS TEMES (2)', 'TOTS ELS TEMES', 'TOTS ELS TEMES (3)']
    
    # Pass 1: Global Answer Map (Text based)
    global_answers = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets: continue
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join([str(c) for c in row if c is not None])
            exam_matches = re.finditer(r'(\d{2}/\d{2}(?:\.\d+)?).*?[:]\s*(.*)', row_text)
            for ex in exam_matches:
                ex_id = ex.group(1)
                qa_text = ex.group(2)
                for qam in re.finditer(r'(\d+)\s*[:.\-)]?\s*([A-D])(?![a-z])', qa_text):
                    global_answers[f"{ex_id}-{qam.group(1)}"] = qam.group(2).lower()
            
            for cell in row:
                cv = str(cell).strip().upper()
                m = re.match(r'^(\d+)\s*[:.\-)]?\s*([A-D])$', cv)
                if m:
                    for ex in ["81/24", "81/23", "81/21", "81/19.2", "81/19.1", "81/16"]:
                        key = f"{ex}-{m.group(1)}"
                        if key not in global_answers: global_answers[key] = m.group(2).lower()

    # Pass 2: Question Extraction + Styling Check
    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets or "DORMIR" in sheet_name.upper(): continue
        print(f"Parsing sheet: {sheet_name}")
        sheet = wb[sheet_name]
        questions = []
        current_exam = "General"
        current_question = None
        
        for row in sheet.iter_rows(values_only=False):
            # Extract row values and styling
            cells = []
            for c in row:
                val = str(c.value).strip() if c.value is not None else ""
                is_styled = False
                if c.font:
                    if c.font.bold or c.font.underline:
                        is_styled = True
                cells.append({"val": val, "styled": is_styled})
            
            anchor_data = None
            for i in [0, 1, 2]: # Look in col A, B or C
                if i < len(cells) and cells[i]["val"]:
                    anchor_data = cells[i]
                    break
            if not anchor_data: continue
            
            anchor = anchor_data["val"]
            
            # Exam header
            ex_h = re.match(r'^(\d{2}/\d{2}(?:\.\d+)?)$', anchor)
            if ex_h:
                if current_question: questions.append(current_question)
                current_exam = ex_h.group(1)
                current_question = None
                continue
            
            # Question start
            q_m = re.match(r'^(\d+)[\s.]+(.*)', anchor)
            if q_m:
                if current_question: questions.append(current_question)
                q_num = q_m.group(1)
                current_question = {
                    "id": f"{sheet_name}-{current_exam}-{q_num}",
                    "exam": current_exam,
                    "num": q_num,
                    "text": q_m.group(2).strip(),
                    "options": {},
                    "answer": global_answers.get(f"{current_exam}-{q_num}", "")
                }
                # Check same-row text answer
                for c_data in cells:
                    am = re.search(r'RESPOSTA[:\s]+([A-D])', c_data["val"].upper())
                    if am: current_question["answer"] = am.group(1).lower()
            
            elif current_question:
                # Option check
                o_m = re.match(r'^([a-d])[\s.)]+\s*(.*)', anchor)
                if o_m:
                    opt_letter = o_m.group(1).lower()
                    opt_text = o_m.group(2).strip()
                    current_question["options"][opt_letter] = opt_text
                    # If this cell is BOLD or UNDERLINED, it's the answer!
                    if anchor_data["styled"]:
                        current_question["answer"] = opt_letter
                elif re.match(r'^\d{2}/\d{2}', anchor):
                    questions.append(current_question)
                    current_question = None
                else:
                    # Multi-line text or option continuation
                    if not current_question["options"]:
                        current_question["text"] += " " + anchor
                    else:
                        lk = list(current_question["options"].keys())[-1]
                        current_question["options"][lk] += " " + anchor

        if current_question: questions.append(current_question)
        if questions:
            valid = [q for q in questions if len(q["options"]) >= 2]
            if valid: data.append({"topic": sheet_name, "questions": valid})
                
    return data

if __name__ == "__main__":
    # Update paths to point to the data directory
    excel_path = 'data/PREGUNTES.xlsx'
    json_output = 'data/questions.json'
    
    res = parse_excel(excel_path)
    with open(json_output, 'w', encoding='utf-8') as f:
        json.dump(res, f, indent=2, ensure_ascii=False)
    print(f"Extraction complete: {len(res)} topics")
