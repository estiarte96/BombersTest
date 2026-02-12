import openpyxl

def check_underline(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    found = False
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Checking sheet: {sheet_name}")
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                if cell.font and cell.font.underline:
                    print(f"  Underline found in {sheet_name} at {cell.coordinate}: {cell.value}")
                    found = True
                    break
            if found: break # Just find one to confirm
        if found: break

if __name__ == "__main__":
    check_underline('PREGUNTES.xlsx')
